from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from .models import PDFHisobot, ImtihonNatija, TalabaYozuv
from .utils import pdf_dan_malumot_ajrat
import zipfile
import os
import re


@login_required
def hisobotlar_list(request):
    hisobotlar = PDFHisobot.objects.filter(
        yuklagan_user=request.user
    ).order_by('-yuklangan_vaqt')
    return render(request, 'hisobotlar/list.html', {'hisobotlar': hisobotlar})


@login_required
def pdf_yuklash(request):
    if request.method == 'POST':
        yuklash_turi = request.POST.get('yuklash_turi')  # 'bitta' yoki 'zip'

        fayllar = []

        if yuklash_turi == 'bitta':
            pdf_fayllar = request.FILES.getlist('pdf_fayllar')
            for fayl in pdf_fayllar:
                if fayl.name.endswith('.pdf'):
                    fayllar.append(fayl)

        elif yuklash_turi == 'zip':
            zip_fayl = request.FILES.get('zip_fayl')
            if zip_fayl and zip_fayl.name.endswith('.zip'):
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp:
                    for chunk in zip_fayl.chunks():
                        tmp.write(chunk)
                    tmp_path = tmp.name

                try:
                    with zipfile.ZipFile(tmp_path, 'r') as z:
                        pdf_nomlar = [n for n in z.namelist() if n.endswith('.pdf')]
                        for pdf_nom in pdf_nomlar:
                            pdf_data = z.read(pdf_nom)
                            from django.core.files.uploadedfile import InMemoryUploadedFile
                            import io
                            pdf_file = InMemoryUploadedFile(
                                io.BytesIO(pdf_data),
                                'file',
                                os.path.basename(pdf_nom),
                                'application/pdf',
                                len(pdf_data),
                                None
                            )
                            fayllar.append(pdf_file)
                finally:
                    os.unlink(tmp_path)

        # Fayllarni qayta ishlash
        muvaffaqiyat = 0
        xato_fayllar = []

        for fayl in fayllar:
            hisobot = PDFHisobot.objects.create(
                fayl=fayl,
                yuklagan_user=request.user,
                holat='qayta_ishlanmoqda'
            )

            # Ma'lumot ajratish
            natija = pdf_dan_malumot_ajrat(hisobot.fayl.path)

            if natija['xato']:
                hisobot.holat = 'xato'
                hisobot.save()
                xato_fayllar.append(fayl.name)
            else:
                # ImtihonNatija — umumiy statistika
                imtihon_natija = ImtihonNatija.objects.create(
                    hisobot=hisobot,
                    fan_nomi=natija['fan_nomi'],
                    fan_oqituvchi=natija['fan_oqituvchi'],
                    guruh=natija['guruh'],
                    jami_talabalar=natija['jami_talabalar'],
                    baho_5=natija['baho_5'],
                    baho_4=natija['baho_4'],
                    baho_3=natija['baho_3'],
                    kelmadi=natija['kelmadi'],
                )

                # TalabaYozuv — har bir talabaning batafsil ma'lumoti
                talaba_yozuvlar = []
                for t in natija.get('talabalar', []):
                    talaba_yozuvlar.append(TalabaYozuv(
                        natija=imtihon_natija,
                        guruh=natija['guruh'],
                        fan_nomi=natija['fan_nomi'],
                        fan_oqituvchi=natija['fan_oqituvchi'],
                        fio=t['fio'],
                        reyting_raqam=t['reyting_raqam'],
                        joriy=t['joriy'],
                        oraliq=t['oraliq'],
                        joriy_oraliq=t['joriy_oraliq'],
                        yakuniy=t['yakuniy'],
                        ozlashtirish=t['ozlashtirish'],
                        baho=t['baho'],
                    ))

                # bulk_create bilan bir marta bazaga yozish (tezroq)
                if talaba_yozuvlar:
                    TalabaYozuv.objects.bulk_create(talaba_yozuvlar)

                hisobot.holat = 'muvaffaqiyatli'
                hisobot.save()
                muvaffaqiyat += 1

        # Natija xabari
        if muvaffaqiyat > 0:
            messages.success(request, f"{muvaffaqiyat} ta PDF muvaffaqiyatli qayta ishlandi!")
        if xato_fayllar:
            messages.warning(
                request,
                f"{len(xato_fayllar)} ta PDF o'qilmadi: {', '.join(xato_fayllar)}"
            )

        return redirect('hisobotlar')

    return render(request, 'hisobotlar/yuklash.html')


@login_required
def natija(request, pk):
    hisobot = get_object_or_404(PDFHisobot, pk=pk, yuklagan_user=request.user)
    natijalar = ImtihonNatija.objects.filter(hisobot=hisobot)
    return render(request, 'hisobotlar/natija.html', {
        'hisobot': hisobot,
        'natijalar': natijalar
    })


@login_required
def excel_yuklab_olish(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import tempfile

    natijalar = ImtihonNatija.objects.filter(
        hisobot__yuklagan_user=request.user
    )
    if pk != 0:
        natijalar = natijalar.filter(hisobot_id=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Imtihon natijalari"

    OQISH   = "FFFFFF"
    KULRANG = "D9D9D9"
    YASHIL  = "C6EFCE"
    SARIQ   = "FFEB9C"
    QIZIL   = "FFC7CE"
    MALLA   = "EDCB9A"
    MOVIY   = "BDD7EE"

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ustunlar = [
        ("T/R",                              OQISH,   6),
        ("Guruh",                            OQISH,  12),
        ("Fan nomi",                         OQISH,  32),
        ("Fan o'qituvchisi",                 OQISH,  36),
        ("Guruhdagi talabalar soni",         KULRANG, 14),
        ('"5"',                              YASHIL,  8),
        ('"4"',                              SARIQ,   8),
        ('"3"',                              QIZIL,   8),
        ("Kelmadi",                          MALLA,   10),
        ("O'zlashtirmagan",                  MOVIY,   14),
        ("O'zlashtirish ko'rsatkichi (%)",   MOVIY,   20),
        ('Sifat ko\'rsatkichi "4","5" (%)',  MOVIY,   20),
    ]

    for col, (nom, rang, kenglik) in enumerate(ustunlar, 1):
        cell = ws.cell(row=1, column=col, value=nom)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor=rang)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = kenglik

    ws.row_dimensions[1].height = 45

    for row, n in enumerate(natijalar, 2):
        E = f"E{row}"
        F = f"F{row}"
        G = f"G{row}"
        H = f"H{row}"

        qatorlar = [
            row - 1,
            n.guruh,
            n.fan_nomi,
            n.fan_oqituvchi,
            n.jami_talabalar,
            n.baho_5,
            n.baho_4,
            n.baho_3,
            n.kelmadi,
            '',
            f"=({F}+{G}+{H})/{E}*100",
            f"=({F}+{G})/{E}*100",
        ]

        for col, qiymat in enumerate(qatorlar, 1):
            cell = ws.cell(row=row, column=col, value=qiymat)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        ws.row_dimensions[row].height = 18

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    with open(tmp_path, 'rb') as f:
        response = HttpResponse(
            f.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="imtihon_natijalari.xlsx"'

    os.unlink(tmp_path)
    return response


@login_required
def pdf_ochirish(request, pk):
    hisobot = get_object_or_404(PDFHisobot, pk=pk, yuklagan_user=request.user)
    if request.method == 'POST':
        if hisobot.fayl and os.path.isfile(hisobot.fayl.path):
            os.remove(hisobot.fayl.path)
        hisobot.delete()
        messages.success(request, "PDF muvaffaqiyatli o'chirildi!")
    return redirect('hisobotlar')


@login_required
def pdf_barchani_ochirish(request):
    if request.method == 'POST':
        hisobotlar = PDFHisobot.objects.filter(yuklagan_user=request.user)
        for h in hisobotlar:
            if h.fayl and os.path.isfile(h.fayl.path):
                os.remove(h.fayl.path)
        hisobotlar.delete()
        messages.success(request, "Barcha PDF lar o'chirildi!")
    return redirect('hisobotlar')