from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse
from datetime import datetime, timedelta
import re


def login_view(request):
    error = None
    if 'login_attempts' not in request.session:
        request.session['login_attempts'] = 0
        request.session['login_blocked_until'] = None
    blocked_until = request.session.get('login_blocked_until')
    if blocked_until:
        blocked_until_dt = datetime.fromisoformat(blocked_until)
        now = datetime.now()
        if now < blocked_until_dt:
            qolgan = int((blocked_until_dt - now).total_seconds())
            return render(request, 'accounts/login.html', {
                'error': f"Juda ko'p urinish! {qolgan} soniyadan so'ng qayta urining.", 'blocked': True})
        else:
            request.session['login_attempts'] = 0
            request.session['login_blocked_until'] = None
    if request.method == 'POST':
        email    = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        blocked_until = request.session.get('login_blocked_until')
        if blocked_until:
            blocked_until_dt = datetime.fromisoformat(blocked_until)
            if datetime.now() < blocked_until_dt:
                qolgan = int((blocked_until_dt - datetime.now()).total_seconds())
                return render(request, 'accounts/login.html', {
                    'error': f"Juda ko'p urinish! {qolgan} soniyadan so'ng qayta urining.", 'blocked': True})
        try:
            user_obj = User.objects.get(email=email)
            user = authenticate(request, username=user_obj.username, password=password)
        except User.DoesNotExist:
            user = None
        if user is not None:
            request.session['login_attempts'] = 0
            request.session['login_blocked_until'] = None
            login(request, user)
            return redirect('dashboard')
        else:
            request.session['login_attempts'] = request.session.get('login_attempts', 0) + 1
            attempts = request.session['login_attempts']
            if attempts >= 3:
                blocked_until = datetime.now() + timedelta(seconds=30)
                request.session['login_blocked_until'] = blocked_until.isoformat()
                request.session['login_attempts'] = 0
                return render(request, 'accounts/login.html', {
                    'error': "3 marta noto'g'ri kiritildi! 30 soniya kuting.", 'blocked': True})
            else:
                error = f"Email yoki parol noto'g'ri! Yana {3-attempts} ta urinish qoldi."
    return render(request, 'accounts/login.html', {'error': error})


def logout_view(request):
    logout(request)
    request.session.flush()
    response = redirect('login')
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required
def filtr_variantlar(request):
    """AJAX: kaskad filtr uchun bog'liq variantlarni qaytaradi"""
    from hisobotlar.models import ImtihonNatija
    fan       = request.GET.get('fan', '').strip()
    oqituvchi = request.GET.get('oqituvchi', '').strip()
    guruh     = request.GET.get('guruh', '').strip()
    qs = ImtihonNatija.objects.filter(hisobot__yuklagan_user=request.user)
    if fan:       qs = qs.filter(fan_nomi=fan)
    if oqituvchi: qs = qs.filter(fan_oqituvchi=oqituvchi)
    if guruh:     qs = qs.filter(guruh=guruh)
    return JsonResponse({
        'fanlar':       sorted(qs.values_list('fan_nomi',      flat=True).distinct()),
        'oqituvchilar': sorted(qs.values_list('fan_oqituvchi', flat=True).distinct()),
        'guruhlar':     sorted(qs.values_list('guruh',         flat=True).distinct()),
    })


@login_required
def dashboard(request):
    from hisobotlar.models import ImtihonNatija, TalabaYozuv
    from django.db.models import Sum

    filter_fan       = request.GET.get('fan',       '').strip()
    filter_oqituvchi = request.GET.get('oqituvchi', '').strip()
    filter_guruh     = request.GET.get('guruh',     '').strip()
    filter_baho_stat = request.GET.get('baho_stat', '').strip()  # 5|4|3|kelmadi

    user_natijalari = ImtihonNatija.objects.filter(hisobot__yuklagan_user=request.user)
    fanlar       = sorted(user_natijalari.values_list('fan_nomi',      flat=True).distinct())
    oqituvchilar = sorted(user_natijalari.values_list('fan_oqituvchi', flat=True).distinct())
    guruhlar     = sorted(user_natijalari.values_list('guruh',         flat=True).distinct())

    # Statistika — filtr bo'lsa filtrlangan, bo'lmasa barcha ma'lumotlar
    natijalar_qs = user_natijalari
    if filter_fan:       natijalar_qs = natijalar_qs.filter(fan_nomi=filter_fan)
    if filter_oqituvchi: natijalar_qs = natijalar_qs.filter(fan_oqituvchi=filter_oqituvchi)
    if filter_guruh:     natijalar_qs = natijalar_qs.filter(guruh=filter_guruh)

    agg = natijalar_qs.aggregate(
        jami_talabalar=Sum('jami_talabalar'),
        jami_5=Sum('baho_5'), jami_4=Sum('baho_4'),
        jami_3=Sum('baho_3'), jami_kelmadi=Sum('kelmadi'),
    )
    jami    = agg['jami_talabalar'] or 0
    besh    = agg['jami_5']         or 0
    tort    = agg['jami_4']         or 0
    uch     = agg['jami_3']         or 0
    kelmadi = agg['jami_kelmadi']   or 0

    statistika = {
        'jami_talabalar': jami,
        'jami_5': besh, 'jami_4': tort, 'jami_3': uch, 'jami_kelmadi': kelmadi,
        'ozlashtirish': round((besh+tort+uch)/jami*100, 1) if jami else 0,
        'sifat':        round((besh+tort)/jami*100, 1)     if jami else 0,
    } if jami > 0 else None

    # Talabalar jadvali — filtr YOKI baho_stat bo'lganda
    filtrlangan_talabalar = None
    filtr_qollandi = bool(filter_fan or filter_oqituvchi or filter_guruh or filter_baho_stat)

    if filtr_qollandi:
        talabalar_qs = TalabaYozuv.objects.filter(natija__hisobot__yuklagan_user=request.user)
        if filter_fan:       talabalar_qs = talabalar_qs.filter(fan_nomi=filter_fan)
        if filter_oqituvchi: talabalar_qs = talabalar_qs.filter(fan_oqituvchi=filter_oqituvchi)
        if filter_guruh:     talabalar_qs = talabalar_qs.filter(guruh=filter_guruh)
        if filter_baho_stat == 'kelmadi':
            talabalar_qs = talabalar_qs.filter(baho=0)
        elif filter_baho_stat in ('5', '4', '3'):
            talabalar_qs = talabalar_qs.filter(baho=int(filter_baho_stat))
        filtrlangan_talabalar = talabalar_qs.order_by('guruh', 'fio')

    return render(request, 'accounts/dashboard.html', {
        'fanlar': fanlar, 'oqituvchilar': oqituvchilar, 'guruhlar': guruhlar,
        'filter_fan': filter_fan, 'filter_oqituvchi': filter_oqituvchi,
        'filter_guruh': filter_guruh, 'filter_baho_stat': filter_baho_stat,
        'filtr_qollandi': filtr_qollandi,
        'statistika': statistika,
        'filtrlangan_talabalar': filtrlangan_talabalar,
    })


@login_required
def umumiy_malumotlar(request):
    from hisobotlar.models import TalabaYozuv, ImtihonNatija

    filter_fan       = request.GET.get('fan', '').strip()
    filter_oqituvchi = request.GET.get('oqituvchi', '').strip()
    filter_guruh     = request.GET.get('guruh', '').strip()
    filter_baho      = request.GET.get('baho', '').strip()
    filter_qidiruv   = request.GET.get('q', '').strip()

    user_natijalari = ImtihonNatija.objects.filter(hisobot__yuklagan_user=request.user)
    fanlar       = sorted(user_natijalari.values_list('fan_nomi',      flat=True).distinct())
    oqituvchilar = sorted(user_natijalari.values_list('fan_oqituvchi', flat=True).distinct())
    guruhlar     = sorted(user_natijalari.values_list('guruh',         flat=True).distinct())

    talabalar_qs = TalabaYozuv.objects.filter(natija__hisobot__yuklagan_user=request.user)
    if filter_fan:       talabalar_qs = talabalar_qs.filter(fan_nomi=filter_fan)
    if filter_oqituvchi: talabalar_qs = talabalar_qs.filter(fan_oqituvchi=filter_oqituvchi)
    if filter_guruh:     talabalar_qs = talabalar_qs.filter(guruh=filter_guruh)
    if filter_baho:      talabalar_qs = talabalar_qs.filter(baho=filter_baho)
    if filter_qidiruv:   talabalar_qs = talabalar_qs.filter(fio__icontains=filter_qidiruv)
    talabalar_qs = talabalar_qs.order_by('guruh', 'fio')

    return render(request, 'accounts/umumiy_malumotlar.html', {
        'talabalar': talabalar_qs, 'jami': talabalar_qs.count(),
        'fanlar': fanlar, 'oqituvchilar': oqituvchilar, 'guruhlar': guruhlar,
        'filter_fan': filter_fan, 'filter_oqituvchi': filter_oqituvchi,
        'filter_guruh': filter_guruh, 'filter_baho': filter_baho, 'filter_qidiruv': filter_qidiruv,
        'filtr_qollandi': bool(filter_fan or filter_oqituvchi or filter_guruh or filter_baho or filter_qidiruv),
    })


@login_required
def dashboard_excel_export(request):
    from hisobotlar.models import TalabaYozuv
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import tempfile, os

    filter_fan       = request.GET.get('fan',       '').strip()
    filter_oqituvchi = request.GET.get('oqituvchi', '').strip()
    filter_guruh     = request.GET.get('guruh',     '').strip()
    filter_baho_stat = request.GET.get('baho_stat', '').strip()

    talabalar_qs = TalabaYozuv.objects.filter(natija__hisobot__yuklagan_user=request.user)
    if filter_fan:       talabalar_qs = talabalar_qs.filter(fan_nomi=filter_fan)
    if filter_oqituvchi: talabalar_qs = talabalar_qs.filter(fan_oqituvchi=filter_oqituvchi)
    if filter_guruh:     talabalar_qs = talabalar_qs.filter(guruh=filter_guruh)
    if filter_baho_stat == 'kelmadi':
        talabalar_qs = talabalar_qs.filter(baho=0)
    elif filter_baho_stat in ('5', '4', '3'):
        talabalar_qs = talabalar_qs.filter(baho=int(filter_baho_stat))
    talabalar_qs = talabalar_qs.order_by('guruh', 'fio')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Talabalar"
    sarlavha_rang = PatternFill("solid", fgColor="BDD7EE")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))
    ustunlar = [
        ("T/R", 6), ("Guruh", 12), ("Fan nomi", 30),
        ("Fan o'qituvchisi", 34), ("Talabaning F.I.Sh", 36),
        ("Reyting daftarcha raqami", 20), ("Joriy (J)", 10),
        ("Oraliq (O)", 10), ("J+O", 10), ("YN", 18),
        ("O'zlashtirish ko'rsatkichi", 22), ("Baho", 8),
    ]
    for col, (nom, kenglik) in enumerate(ustunlar, 1):
        cell = ws.cell(row=1, column=col, value=nom)
        cell.font = Font(bold=True, size=10)
        cell.fill = sarlavha_rang
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = kenglik
    ws.row_dimensions[1].height = 36
    baho_ranglar = {5: "C6EFCE", 4: "FFEB9C", 3: "FFC7CE"}
    for row_idx, t in enumerate(talabalar_qs, 2):
        for col, qiymat in enumerate([
            row_idx-1, t.guruh, t.fan_nomi, t.fan_oqituvchi,
            t.fio, t.reyting_raqam, t.joriy, t.oraliq,
            t.joriy_oraliq, t.yakuniy, t.ozlashtirish, t.baho,
        ], 1):
            cell = ws.cell(row=row_idx, column=col, value=qiymat)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if col == 12 and t.baho in baho_ranglar:
                cell.fill = PatternFill("solid", fgColor=baho_ranglar[t.baho])
        ws.row_dimensions[row_idx].height = 16

    qismlar = []
    if filter_fan:       qismlar.append(filter_fan[:15])
    if filter_guruh:     qismlar.append(filter_guruh)
    if filter_baho_stat: qismlar.append(f'baho_{filter_baho_stat}')
    fayl_nomi = re.sub(r"[^\w\-.]", "_", ('_'.join(qismlar) or 'talabalar') + '.xlsx')

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name); tmp_path = tmp.name
    with open(tmp_path, 'rb') as f:
        response = HttpResponse(f.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{fayl_nomi}"'
    os.unlink(tmp_path)
    return response


@login_required
def sozlamalar(request):
    message = None; error = None
    if request.method == 'POST':
        eski_parol   = request.POST.get('eski_parol')
        yangi_parol1 = request.POST.get('yangi_parol1')
        yangi_parol2 = request.POST.get('yangi_parol2')
        if not request.user.check_password(eski_parol):
            error = "Eski parol noto'g'ri!"
        elif yangi_parol1 != yangi_parol2:
            error = "Yangi parollar mos kelmadi!"
        elif len(yangi_parol1) < 8:
            error = "Parol kamida 8 ta belgidan iborat bo'lishi kerak!"
        elif not re.search(r'[A-Z]', yangi_parol1):
            error = "Parolda kamida 1 ta katta harf bo'lishi kerak (A-Z)!"
        elif not re.search(r'[0-9]', yangi_parol1):
            error = "Parolda kamida 1 ta raqam bo'lishi kerak (0-9)!"
        else:
            request.user.set_password(yangi_parol1)
            request.user.save()
            update_session_auth_hash(request, request.user)
            message = "Parol muvaffaqiyatli yangilandi!"
    return render(request, 'accounts/sozlamalar.html', {'message': message, 'error': error})


@login_required
def session_yangilash(request):
    request.session.modified = True
    return JsonResponse({'status': 'ok'})
